from openpyxl import load_workbook

#   STANDART FUNKSIYALAR
# Excel database'ga ulanish
def dbxl_connect():
    wb = load_workbook('database.xlsx')
    sheet = wb.active
    return sheet


# Ma'lumotni olish, cordinate orqali
def get_data(cordinate):
    sheet = dbxl_connect()
    data = sheet[cordinate].value
    return data


# Ma'lumot qo'shish uchun, cordinata va ma'lumot kiritish orqali
def add_data(cordinate, item):
    wb = load_workbook('database.xlsx')
    sheet = wb.active
    sheet[cordinate] = item
    wb.save('database.xlsx')



#   QO'SHIMCHA FUNKSIYALAR
# Foydalanuvchi qo'shish uchun. id, fullname kiritish orqali
def add_user(id, fullname, username = None):
    id = str(id)

    wb = load_workbook('database.xlsx') 
    sheet = wb.active
    
    IDs = []
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    if id not in IDs:
        N = len(IDs) + 1
        sheet[f'A{N}'] = id
        sheet[f'B{N}'] = fullname
        if username != None:
            sheet[f'C{N}'] = f"@{username}"
        wb.save('database.xlsx')


# 1-ustunda foydalanuvchi id'si bo'lsa, u nechanchi qatorda ekanligini qaytaradi 
def get_num_user(id):
    sheet = dbxl_connect()
    IDs = []
    
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    if id not in IDs:
        N = len(IDs)
    else:
        N = IDs.index(id) + 1
    
    return N


def getWalletData(id, walletName):
    
    wb = load_workbook(filename = "database.xlsx")
    sheet = wb.active
    IDs = []    
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    N=1            
    for ID in IDs: 
        if str(id) == ID:
            break
        N+=1
    
    wallet_dict = {'uzcard':'G', 'humo':'I', 'qiwi':'K', 'wmz':'L', 'wmr':'M', 'payeer':'N', 'yandex':'O', 'sberbank':'P', 'tinkoff':'Q', 'payeer_rub':'N', 'payeer_usd':'N'}
    for key in wallet_dict:
        if walletName == key:
            return sheet[f'{wallet_dict[key]}{N}'].value


### Add UZCARD
def addWallet_db(message, let):
        wb = load_workbook(filename = "database.xlsx")
        sheet = wb.active
        IDs = []    
        for row in sheet.rows:
            IDs.append(str(row[0].value))
    
        if str(message.from_user.id) in IDs:
            N = IDs.index(str(message.from_user.id)) + 1       
        else:
            N = sheet.max_row
        
        sheet[f'{let}{N}'] = message.text
        wb.save('database.xlsx')

def get_cordinate(N, item):
    let_list = ['D', 'E', 'F', 'J', 'K', 'L', 'M', 'N']
    sheet = dbxl_connect()
    
    for let in let_list:
        if sheet[f"{let}{N}"].value == item:
            return f"{let}{N}"
    
def clearWallets(call):
    wb = load_workbook(filename = "database.xlsx")
    sheet = wb.active
    IDs = []    
    for row in sheet.rows:
        IDs.append(str(row[0].value))
  
    if str(call.from_user.id) in IDs:
        N = IDs.index(str(call.from_user.id)) + 1       
    else:
        N = sheet.max_row
    
    for let in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
        sheet[f'{let}{N}'] = None
        
    wb.save('database.xlsx')
    

def getTransferId():
    wb = load_workbook('transfers.xlsx') 
    sheet = wb.active
    
    IDs = []
    for row in sheet.rows:
        IDs.append(str(row[0].value))

    N = len(IDs) + 1
    return N

def addOperation(berish, olish, berish_num, olish_num, berish_miqdor, olish_miqdor, fullname, username, user_id):
    wb = load_workbook('transfers.xlsx') 
    sheet = wb.active
    
    IDs = []
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    N = len(IDs) + 1
    sheet[f'A{N}'] = N-1
    sheet[f'B{N}'] = berish
    sheet[f'C{N}'] = olish
    sheet[f'D{N}'] = berish_num
    sheet[f'E{N}'] = olish_num
    sheet[f'F{N}'] = berish_miqdor
    sheet[f'G{N}'] = olish_miqdor
    sheet[f'H{N}'] = fullname
    sheet[f'J{N}'] = user_id
    
            
    if username != None:
        sheet[f'I{N}'] = f"@{username}"
    
    wb.save('transfers.xlsx')
    
# Transferni holatini belgilaydi va user_id ni qaytaradi
def check_trans(trans_id, status):
    wb = load_workbook('transfers.xlsx') 
    sheet = wb.active
    IDs = []
    
    for row in sheet.rows:
        IDs.append(str(row[0].value))
    
    # if trans_id not in IDs:
    #     N = len(IDs)
    #     print("Yoq")
    # else:
    N = IDs.index(str(int(trans_id)-1)) + 1
        # print("Ha")
    sheet[f'K{N}'] = status
    user_id = sheet[f'J{N}'].value
    wb.save('transfers.xlsx')
    
    return user_id
    
    